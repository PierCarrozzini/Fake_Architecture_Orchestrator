import xml.etree.ElementTree as ET
from tabulate import tabulate
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_drawio_xml(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    default_value1 = 80
    default_value2 = 8080
    default_value3 = ''
    default_value4 = 'yes'

    components = []
    component_counter = 1
    recognized_values = ['web_server', 'database', 'firewall', 'firewall-alpine', 'cms',
                         'cache', 'message_queque', 'proxy', 'monitoring', 'prometheus', 'grafana',
                         'ci_cd', 'e_commerce', 'machine_learning', 'version_control']
    value_counts = {}
    other_value_count = 0

    for cell_elem in root.findall('.//'):
        if cell_elem.tag == 'object':
            component_info = {}

            component_info['number'] = component_counter
            component_counter += 1

            # Extract information about each vertex (component)
            component_info['id'] = cell_elem.attrib.get('id')
            component_info['name'] = cell_elem.attrib.get('id')
            component_info['value'] = cell_elem.attrib.get('label')
            component_info['type'] = cell_elem.attrib.get('label')
            component_info['Internal_port'] = cell_elem.attrib.get('Internal_port') if cell_elem.attrib.get(
                'Internal_port') is not None else default_value1
            component_info['External_port'] = cell_elem.attrib.get('External_port') if cell_elem.attrib.get(
                'External_port') is not None else default_value2

            sottoelem = cell_elem.find('.//mxCell')
            if sottoelem is not None:
                component_info['style'] = sottoelem.get('style')
                geometry_elem = cell_elem.find('.//mxGeometry')
                if geometry_elem is not None:
                    component_info['x'] = geometry_elem.get('x')
                    component_info['y'] = geometry_elem.get('y')
                    component_info['width'] = geometry_elem.get('width')
                    component_info['height'] = geometry_elem.get('height')

            components.append(component_info)

            data_string = component_info
            key_value_table = [(key, str(value)) for key, value in data_string.items()]
            print(tabulate(key_value_table, headers=["Chiave", "Valore"], tablefmt="orgtbl"))

            print(f"Component :")
            print(f"Number: {component_info['number']}")
            print(f"ID: {component_info['id']}")
            print(f"Name: {component_info['id']}")
            print(f"Value: {component_info['value']}")
            print(f"Type: {component_info['value']}")
            print(f"Style: {component_info['style']}")
            print(f"Internal_port: {component_info['Internal_port']}")
            print(f"External_port: {component_info['External_port']}")
            print(f"Position: ({component_info['x']}, {component_info['y']})")
            print(f"Size: {component_info['width']} x {component_info['height']}")
            print()

            # Aggiunta della logica di conteggio
            value = component_info['value']
            if value in recognized_values or recognized_values:
                if value in value_counts:
                    value_counts[value] += 1
                    print(f"Found {value} component in the diagram, number = {value_counts[value]}.")
                else:
                    value_counts[value] = 1
                    print(f"Found {value} component in the diagram, number = {value_counts[value]}.")
            else:
                other_value_count += 1
                print(f"Found a not recognized component in the diagram, number = {other_value_count}.")
                print(f"The unrecognized component is named '{value}'.")

        elif cell_elem.tag == 'mxCell' and cell_elem.get('value') is not None:

            component_info = {}

            component_info['number'] = component_counter
            component_counter += 1

            # Extract information about each vertex (component)
            component_info['id'] = cell_elem.get('id')
            component_info['name'] = cell_elem.get('id')
            component_info['value'] = cell_elem.get('value')
            component_info['type'] = cell_elem.get('value')
            component_info['Internal_port'] = cell_elem.attrib.get('Internal_port') if cell_elem.attrib.get(
                'Internal_port') is not None else default_value3
            component_info['External_port'] = cell_elem.attrib.get('External_port') if cell_elem.attrib.get(
                'External_port') is not None else default_value3
            component_info['style'] = cell_elem.get('style')

            # Extract geometry information
            geometry_elem = cell_elem.find('mxGeometry')
            if geometry_elem is not None:
                component_info['x'] = geometry_elem.get('x')
                component_info['y'] = geometry_elem.get('y')
                component_info['width'] = geometry_elem.get('width')
                component_info['height'] = geometry_elem.get('height')
            if component_info['style'] is not None:
                components.append(component_info)

                data_string = component_info
                key_value_table = [(key, str(value)) for key, value in data_string.items()]
                print(tabulate(key_value_table, headers=["Chiave", "Valore"], tablefmt="orgtbl"))

                print(f"Component :")
                print(f"Number: {component_info['number']}")
                print(f"ID: {component_info['id']}")
                print(f"Name: {component_info['id']}")
                print(f"Value: {component_info['value']}")
                print(f"Type: {component_info['value']}")
                print(f"Style: {component_info['style']}")
                print(f"Position: ({component_info['x']}, {component_info['y']})")
                print(f"Size: {component_info['width']} x {component_info['height']}")
                print(f"Internal_Port: {component_info['Internal_port']}")
                print(f"External_Port: {component_info['External_port']}")
                # mancano i print di int/ext port perché è tardi
                print()

                # Aggiunta della logica di conteggio
                value = component_info['value']
                if value in recognized_values:
                    if value in value_counts:
                        value_counts[value] += 1
                        print(f"Found {value} component in the diagram, number = {value_counts[value]}.")
                    else:
                        value_counts[value] = 1
                        print(f"Found {value} component in the diagram, number = {value_counts[value]}.")
                else:
                    other_value_count += 1
                    print(f"Found a not recognized component in the diagram, number = {other_value_count}.")
                    print(f"The unrecognized component is named '{value}'.")

        print('---------------------------------------------')

    print(f"Found {len(components)} components in the diagram.")

    # Creating individual variables for each unique value
    for value, count in value_counts.items():
        if value in recognized_values:
            globals()[value] = count
            print(f"Number of '{value}' components: {count}")
    print(f"Number of not recognized components: {other_value_count}")
    print(components)
    # Tabella
    data = components

    # Ottiengo le chiavi del dizionario come header (per tabella è così)
    header = list(data[0].keys())
    print(header)

    # Conversione dei dizionari in liste per ogni riga
    table_data = [list(row.values()) for row in data]
    print(table_data)

    # Stampo i dati sotto forma di tabella (package tabulate)
    print(tabulate(table_data, headers=header, tablefmt="fancy_grid"))

    excel_filename = "Report/output.xlsx"

    next_sheet_number = 1
    while f'Foglio{next_sheet_number}' in pd.ExcelFile(excel_filename).sheet_names:
        next_sheet_number += 1

    df = pd.DataFrame(table_data,
                      columns=['number', 'id', 'name', 'value', 'type', 'Internal_port', 'External_port', 'style', 'x',
                               'y', 'width', 'height'])

    with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:

        if not pd.ExcelFile(excel_filename).sheet_names:
            writer.book.create_sheet(title=f'Foglio{next_sheet_number}')
            sheet = writer.sheets[f'Foglio{next_sheet_number}']

            sheet.append([None])
            sheet.append([None])

        df.to_excel(writer, index=False, sheet_name=f'Foglio{next_sheet_number}', startrow=4)  # Inizia dalla riga 4

        separator_row = len(df) + 8
        sheet = writer.sheets[f'Foglio{next_sheet_number}']
        sheet.insert_rows(separator_row)
        sheet.cell(row=separator_row, column=1, value='#----------------------------------------------#')
        sheet.cell(row=2, column=8, value="*Fake Architecture Orchestrator* Run Number=")
        sheet.cell(row=2, column=9, value=next_sheet_number)
        sheet.column_dimensions[get_column_letter(8)].width = 40

        print(f"I dati sono stati aggiunti con successo (Foglio{next_sheet_number}).")

    return components
